"""Formato compartido por los exportadores a Excel de los análisis de medidores.

Los colores de serie salen de una paleta categórica validada para daltonismo (ΔE 24.7 en
protanopia entre las dos series). Los colores de estado son un juego aparte y reservado: nunca
se usan como color de serie, y siempre acompañan a un texto ("CUMPLE" / "SIN SEÑAL" / …), de
modo que ningún dato dependa del color solo para leerse.
"""
from datetime import datetime, timedelta, timezone

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colombia: UTC-5 fijo, sin horario de verano (convención del repo).
BOGOTA = timezone(timedelta(hours=-5))

AZUL = "2A78D6"      # serie 1
NARANJA = "EB6834"   # serie 2
TINTA = "0B0B0B"
TINTA_2 = "52514E"
GRIS_SUAVE = "F2F2F0"

# Rellenos de estado (fondo claro + texto oscuro del mismo tono, para que contraste).
OK_BG, OK_FG = "DCF2DC", "076B07"
MAL_BG, MAL_FG = "F8DEDE", "8E2020"
NA_BG, NA_FG = "FDF1D6", "8A6100"

BORDE_FINO = Border(bottom=Side(style="thin", color="D8D8D4"))


def a_bogota(ts_iso):
    return datetime.fromisoformat(ts_iso.replace("Z", "+00:00")).astimezone(BOGOTA).replace(tzinfo=None)


def titulo(ws, fila, texto, tam=14):
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(bold=True, size=tam, color=TINTA)
    return fila + 1


def nota(ws, fila, texto, cols=1):
    """Nota al pie. Con cols>1 se fusiona y ajusta, para que no la corte la celda de al lado."""
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(size=10, color=TINTA_2, italic=True)
    if cols > 1:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=cols)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ancho = sum(ws.column_dimensions[get_column_letter(i)].width or 10 for i in range(1, cols + 1))
        ws.row_dimensions[fila].height = 14 * max(1, -(-len(texto) // max(20, int(ancho * 1.05))))
    return fila + 1


def encabezados(ws, fila, cols, alto=30):
    for i, nombre in enumerate(cols, start=1):
        c = ws.cell(row=fila, column=i, value=nombre)
        c.font = Font(bold=True, size=10, color=TINTA)
        c.fill = PatternFill("solid", fgColor=GRIS_SUAVE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE_FINO
    ws.row_dimensions[fila].height = alto
    return fila + 1


def anchos(ws, medidas):
    for i, w in enumerate(medidas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def pinta_estado(celda, ok, tam=10):
    """ok: True / False / None (sin señal). El texto de la celda ya dice cuál es; el color
    solo refuerza — nunca es el único portador del significado."""
    if ok is True:
        celda.fill = PatternFill("solid", fgColor=OK_BG)
        celda.font = Font(bold=True, size=tam, color=OK_FG)
    elif ok is False:
        celda.fill = PatternFill("solid", fgColor=MAL_BG)
        celda.font = Font(bold=True, size=tam, color=MAL_FG)
    else:
        celda.fill = PatternFill("solid", fgColor=NA_BG)
        celda.font = Font(bold=True, size=tam, color=NA_FG)


def pct(v):
    """Los analizadores entregan porcentajes en unidades de %; Excel los quiere como fracción."""
    return None if v is None else v / 100.0


def estiliza_ejes(g, tick_skip=180):
    """Ejes y rejilla recesivos, etiquetas abajo, leyenda fuera del área de trazado.

    `tickLblPos = "low"` no es opcional: sin eso Excel ancla las etiquetas del eje X en y=0, y
    con series negativas (Gecelca mide en frontera de entrada) el cero queda por encima del
    trazo y las etiquetas se imprimen sobre los datos.
    """
    g.x_axis.delete = False
    g.y_axis.delete = False
    g.x_axis.majorTickMark = "none"
    g.y_axis.majorTickMark = "none"
    g.x_axis.tickLblSkip = tick_skip
    g.x_axis.tickMarkSkip = tick_skip
    g.x_axis.tickLblPos = "low"
    g.y_axis.tickLblPos = "low"
    if g.legend is not None:
        g.legend.position = "b"
        g.legend.overlay = False
    g.style = 2


def linea(serie, color, ancho_pt=1):
    """Marca fina, sin suavizado: son series de cientos de puntos."""
    serie.graphicalProperties.line.solidFill = color
    serie.graphicalProperties.line.width = int(12700 * ancho_pt)
    serie.smooth = False
    return serie
