#!/usr/bin/env python3
"""Exporta a Excel la comparación entre los dos bancos de registros de energía del ION8650.

El medidor publica las mismas magnitudes dos veces, con IDs y codificación distintos:
    Banco A — 40230..40238, INT32 clásico (base 65536).
    Banco B — 40091..40105, INT32-M10K (base 10000).
Este archivo pone cada magnitud lado a lado con su contraparte y dice cuál ID usar.

Las cifras no se recalculan acá: salen de `node scripts/analyze-registros.js --json`, el mismo
código que produce el reporte de consola. Este script da formato y arma el detalle tick a tick
desde los JSONL crudos.

Uso (desde server/):
    python scripts/export-registros-xlsx.py [salida.xlsx]

Requiere `openpyxl` y `node` en el PATH.
"""
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font

sys.path.insert(0, str(Path(__file__).resolve().parent / "lib"))
from xlsxfmt import (  # noqa: E402
    AZUL, BOGOTA, BORDE_FINO, NARANJA, TINTA, TINTA_2,
    a_bogota, anchos, encabezados, estiliza_ejes, linea, nota, pinta_estado, titulo,
)

AQUI = Path(__file__).resolve().parent
TRACES = AQUI.parent / "traces" / "registros"
ANALIZADOR = AQUI / "analyze-registros.js"


def cargar_resumen():
    res = subprocess.run(
        ["node", str(ANALIZADOR), "--json"],
        capture_output=True, text=True, encoding="utf-8", cwd=str(AQUI.parent),
    )
    if res.returncode != 0:
        sys.exit(f"El analizador falló:\n{res.stderr}")
    return json.loads(res.stdout)


def cargar_jsonl():
    if not TRACES.is_dir():
        sys.exit(f"No existe {TRACES}. Corre primero: npm run shadow:registros")
    por_medidor = {}
    for archivo in sorted(TRACES.glob("registros-*.jsonl")):
        with archivo.open(encoding="utf-8") as fh:
            for linea_txt in fh:
                linea_txt = linea_txt.strip()
                if not linea_txt:
                    continue
                r = json.loads(linea_txt)
                por_medidor.setdefault(r["medidor"], []).append(r)
    for v in por_medidor.values():
        v.sort(key=lambda r: r["ts"])
    return por_medidor


# ─── Hojas ────────────────────────────────────────────────────────────────────

def hoja_veredicto(wb, data):
    ws = wb.create_sheet("Veredicto")
    anchos(ws, [16, 11, 11, 24, 11, 62])
    v = data["ventana"]
    f = titulo(ws, 1, "¿Cuál de los IDs duplicados sirve?", 15)
    f = nota(ws, f, f"El ION8650 publica las mismas magnitudes en dos bancos. Ventana "
                    f"{a_bogota(v['desde']):%Y-%m-%d %H:%M} → {a_bogota(v['hasta']):%H:%M} hora Bogotá "
                    f"({v['horas']:.2f} h), {data['registros']} lecturas sobre "
                    f"{len(data['medidores'])} medidores · generado {datetime.now(BOGOTA):%Y-%m-%d %H:%M}", cols=6)
    f += 1

    f = titulo(ws, f, "Magnitudes que existen en los dos bancos", 12)
    f = encabezados(ws, f, ["Magnitud", "Banco A\n(INT32)", "Banco B\n(Mod10K)",
                            "¿Publican el mismo valor?", "Usar", "Por qué"], alto=34)
    for g in data["global"]:
        problemas = []
        for lado, reg in (("A", g["a"]), ("B", g["b"])):
            if reg["problemas"]:
                medidores = ", ".join(sorted({p["medidor"] for p in reg["problemas"]}))
                motivo = reg["problemas"][0]["motivo"]
                problemas.append(f"{reg['reg']} (banco {lado}) falla en {medidores}: {motivo}")
        razon = " · ".join(problemas) if problemas else "los dos registros están sanos en los 6 medidores; se prefiere el banco A por no tener tope de formato"
        vals = [g["magnitud"], g["a"]["reg"], g["b"]["reg"],
                "sí, valor idéntico" if g["identicoEnTodos"] else "NO — ver detalle",
                g["recomendado"], razon]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=f, column=i, value=val)
            c.border = BORDE_FINO
            c.font = Font(size=10, color=TINTA, bold=(i in (1, 5)))
            c.alignment = Alignment(vertical="top", wrap_text=(i == 6))
            if i == 4:
                pinta_estado(c, g["identicoEnTodos"])
            if i == 5:
                pinta_estado(c, True)
        ws.row_dimensions[f].height = 14 * max(2, -(-len(razon) // 62))
        f += 1

    f += 1
    f = titulo(ws, f, "Registros sin contraparte (existen en un solo banco)", 12)
    f = encabezados(ws, f, ["Magnitud", "Registro", "Banco", "Estado", "", ""])
    for r in data["sinContraparte"]:
        malos = r["inservibleEn"]
        vals = [r["magnitud"], r["reg"], r["banco"],
                "sano en todos" if not malos else f"inservible en {len(malos)}/{len(data['medidores'])}: {', '.join(malos)}", "", ""]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=f, column=i, value=val)
            c.border = BORDE_FINO
            c.font = Font(size=10, color=TINTA)
            if i == 4:
                pinta_estado(c, not malos)
                c.alignment = Alignment(vertical="top", wrap_text=True)
        f += 1

    f += 1
    f = titulo(ws, f, "Lo que hay que saber", 12)
    for txt in [
        "· Los dos bancos llevan EXACTAMENTE el mismo número en las 4 magnitudes que comparten. No son "
        "mediciones distintas: es el mismo acumulador publicado con dos codificaciones.",
        "· El banco B usa Mod10K: cada registro de 16 bits lleva 4 dígitos DECIMALES, así que el valor es "
        "w0·10000 + w1, no w0·65536 + w1. Leerlo como INT32 da un número plausible pero equivocado — y "
        "traicionero, porque mientras w0 no cambie el INCREMENTO coincide y solo se rompe cuando w1 da la vuelta.",
        "· El banco B tiene techo: 327.679.999. Cuando el acumulador lo pasa, el registro se queda clavado en "
        "el centinela 0x7FFF/0 y su Δ da cero para siempre, como si el contador estuviera congelado.",
        "· Los registros del banco B que combinan (del+rec, del−rec) dan la vuelta a los 10.000.000, así que "
        "su valor absoluto no sirve; hay que calcular la combinación a partir de del y rec.",
        "· Por eso la recomendación es el banco A (40230..40238): mismo dato, sin techo de formato y con la "
        "codificación INT32 que ya usa el resto del proyecto.",
    ]:
        c = ws.cell(row=f, column=1, value=txt)
        c.font = Font(size=10, color=TINTA)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=6)
        ws.row_dimensions[f].height = 14 * max(2, -(-len(txt) // 118))
        f += 1
    return ws


def hoja_lado_a_lado(wb, data):
    ws = wb.create_sheet("Lado a lado")
    anchos(ws, [15, 14, 11, 11, 17, 17, 15, 26])
    f = titulo(ws, 1, "Cada magnitud contra su contraparte, medidor por medidor")
    f = nota(ws, f, "«Δ ventana» es cuánto avanzó el registro durante la hora. Si los dos bancos publican el "
                    "mismo acumulador, los dos Δ tienen que coincidir — y cuando uno de los dos no coincide, "
                    "la columna «Estado» dice por qué.", cols=8)
    f += 1
    f = encabezados(ws, f, ["Medidor", "Magnitud", "Reg. A", "Reg. B", "Δ ventana A", "Δ ventana B",
                            "¿Mismo valor?", "Estado"])
    primera = f
    for m in data["medidores"]:
        for p in m["pares"]:
            usables = p["iguales"] + p["porDesfase"] + p["reales"]
            if usables == 0:
                mismo, ok = "no comparable", None
            elif p["reales"] == 0:
                mismo, ok = (f"sí, {p['iguales']}/{p['iguales']}" if p["porDesfase"] == 0
                             else f"sí ({p['porDesfase']} por desfase)"), True
            else:
                mismo, ok = f"NO: {p['reales']} difieren", False
            estado = p["inservibleB"] or p["inservibleA"] or "ambos sanos"
            vals = [m["medidor"], p["magnitud"], p["regA"], p["regB"], p["deltaA"], p["deltaB"], mismo, estado]
            for i, val in enumerate(vals, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.border = BORDE_FINO
                c.font = Font(size=10, color=TINTA)
                c.alignment = Alignment(vertical="top", wrap_text=(i == 8))
                if i in (5, 6):
                    c.number_format = "#,##0"
                if i == 7:
                    pinta_estado(c, ok)
                if i == 8 and estado != "ambos sanos":
                    pinta_estado(c, False)
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)
    ws.auto_filter.ref = f"A{primera - 1}:H{f - 1}"
    return ws


def hoja_salud(wb, data):
    ws = wb.create_sheet("Salud registros")
    anchos(ws, [15, 10, 6, 16, 15, 15, 15, 11, 11, 12, 46])
    f = titulo(ws, 1, "Salud de cada registro en cada medidor")
    f = nota(ws, f, "Una lectura saturada o malformada NO es una medida, así que no cuenta para el Δ de la "
                    "ventana: el Δ se calcula entre la primera y la última lectura utilizable.", cols=11)
    f += 1
    f = encabezados(ws, f, ["Medidor", "Registro", "Banco", "Magnitud", "Valor inicial", "Valor final",
                            "Δ ventana", "Saturado", "Malformado", "Retrocesos", "Observación"])
    primera = f
    for m in data["medidores"]:
        for s in m["salud"]:
            obs = s["inservible"] or ("no se movió en toda la ventana" if s["congelado"]
                                      else "sube de forma monótona, sin saturar")
            vals = [m["medidor"], s["reg"], s["banco"], s["magnitud"], s["valorInicial"], s["valorFinal"],
                    s["delta"], s["ticksSat"], s["ticksMal"], s["retrocesos"], obs]
            for i, val in enumerate(vals, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.border = BORDE_FINO
                c.font = Font(size=9, color=TINTA)
                c.alignment = Alignment(vertical="top", wrap_text=(i == 11))
                if i in (5, 6, 7):
                    c.number_format = "#,##0"
                if i == 11:
                    pinta_estado(c, None if s["inservible"] else (False if s["congelado"] else True), tam=9)
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)
    ws.auto_filter.ref = f"A{primera - 1}:K{f - 1}"
    return ws


def hoja_identidades(wb, data):
    ws = wb.create_sheet("Identidades banco B")
    anchos(ws, [15, 30, 11, 11, 20, 62])
    f = titulo(ws, 1, "Identidades internas del banco B")
    f = nota(ws, f, "Si la decodificación Mod10K es correcta, «del+rec» tiene que ser exactamente «del» más "
                    "«rec». Es la mejor verificación posible porque no depende de ninguna referencia externa: "
                    "compara el banco contra sí mismo.", cols=6)
    f += 1
    f = encabezados(ws, f, ["Medidor", "Identidad", "Se cumple", "Falla", "Resultado", "Modo de falla"])
    primera = f
    for m in data["medidores"]:
        for i2 in m["identidades"]:
            total = i2["ok"] + i2["falla"]
            if total == 0:
                res, ok = "no verificable", None
            elif i2["cumple"]:
                res, ok = "se cumple siempre", True
            elif i2.get("modoFalla") == "redondeo":
                res, ok = "difiere solo por redondeo", None
            else:
                res, ok = "no se cumple", False
            vals = [m["medidor"], i2["nombre"], i2["ok"], i2["falla"], res, i2.get("explicacion") or ""]
            for j, val in enumerate(vals, start=1):
                c = ws.cell(row=f, column=j, value=val)
                c.border = BORDE_FINO
                c.font = Font(size=10, color=TINTA)
                c.alignment = Alignment(vertical="top", wrap_text=(j == 6))
                if j == 5:
                    pinta_estado(c, ok)
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)
    ws.auto_filter.ref = f"A{primera - 1}:F{f - 1}"
    return ws


def hoja_energia_real(wb, data):
    ws = wb.create_sheet("Contraste energía real")
    anchos(ws, [15, 11, 6, 16, 17, 20, 17, 15])
    f = titulo(ws, 1, "Contraste contra la energía que hubo de verdad")
    f = nota(ws, f, "Referencia independiente de los acumuladores: se integra la potencia del registro 40204 "
                    "(kW tot) sobre la ventana. El registro que sirve es el que reproduce esa energía. En "
                    "Gecelca la frontera es de entrada, así que con la planta generando avanza «rec» y «del» "
                    "se queda quieto.", cols=8)
    f += 1
    f = encabezados(ws, f, ["Medidor", "Registro", "Banco", "Magnitud", "Δ ventana",
                            "Energía real ∫|P|dt (kWh)", "Δ / energía real", "Error"])
    primera = f
    for m in data["medidores"]:
        for c2 in m["contraste"]:
            vals = [m["medidor"], c2["reg"], c2["banco"], c2["magnitud"], c2["delta"],
                    m["integralKwh"], c2["factor"], (c2["errPct"] / 100.0) if c2["errPct"] is not None else None]
            for i, val in enumerate(vals, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.border = BORDE_FINO
                c.font = Font(size=10, color=TINTA)
                if i in (5, 6):
                    c.number_format = "#,##0"
                elif i == 7:
                    c.number_format = "0.0000"
                elif i == 8:
                    c.number_format = "0.00%"
                    pinta_estado(c, abs(c2["errPct"]) <= 2.0 if c2["errPct"] is not None else None)
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)
    ws.auto_filter.ref = f"A{primera - 1}:H{f - 1}"
    return ws


def hoja_detalle(wb, data, por_medidor):
    """Una fila por lectura, con los 13 registros y las 4 diferencias A−B lado a lado."""
    ws = wb.create_sheet("Detalle")
    cat = data["catalogo"]
    regs = cat["bancoA"] + cat["bancoB"]
    pares = cat["pares"]

    cols = ["Hora (Bogotá)", "Medidor", "kW tot", "Orden de lectura"]
    cols += [f"{r['reg']} {r['magnitud']} ({r['banco']})" for r in regs]
    cols += [f"A−B {p['magnitud']}" for p in pares]
    anchos(ws, [17, 15, 13, 15] + [17] * len(regs) + [15] * len(pares))

    f = titulo(ws, 1, "Detalle tick a tick — los dos bancos leídos en el mismo instante")
    f = nota(ws, f, "«Orden de lectura» alterna A→B y B→A entre ticks a propósito: los dos bancos están en "
                    "bloques Modbus distintos y se leen con unos ms de diferencia, durante los cuales el "
                    "contador sigue subiendo. Si una diferencia A−B viene de ese desfase, cambia de signo "
                    "cuando cambia el orden; si fuera un desacuerdo real entre registros, no.", cols=8)
    f += 1
    cab = f
    f = encabezados(ws, f, cols)
    inicio = f

    for medidor in sorted(por_medidor):
        for r in por_medidor[medidor]:
            if not r.get("ok"):
                continue
            fila = [a_bogota(r["ts"]), medidor, r.get("kwTot"), r.get("ordenLectura")]
            for reg in regs:
                v = r["regs"].get(str(reg["reg"])) or {}
                # Un centinela de saturación no es una medida: se escribe el texto en vez del
                # número, para que nadie lo promedie ni lo grafique por error.
                fila.append("SATURADO" if v.get("sat") else ("FORMATO INVÁLIDO" if v.get("mal") else v.get("value")))
            for p in pares:
                a = r["regs"].get(str(p["a"])) or {}
                b = r["regs"].get(str(p["b"])) or {}
                usable = (a.get("value") is not None and b.get("value") is not None
                          and not a.get("sat") and not a.get("mal") and not b.get("sat") and not b.get("mal"))
                fila.append(a["value"] - b["value"] if usable else None)
            for i, val in enumerate(fila, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.font = Font(size=9, color=TINTA)
                if i == 1:
                    c.number_format = "hh:mm:ss"
                elif i == 3:
                    c.number_format = "#,##0.000"
                elif i >= 5:
                    c.number_format = "#,##0"
                if isinstance(val, str) and val in ("SATURADO", "FORMATO INVÁLIDO"):
                    pinta_estado(c, False, tam=9)
            f += 1

    ws.freeze_panes = ws.cell(row=inicio, column=1)
    ws.auto_filter.ref = f"A{cab}:{ws.cell(row=cab, column=len(cols)).coordinate[:-len(str(cab))]}{f - 1}"
    return ws


def rango_redondeado(lo, hi, margen_rel=0.12, divisiones=6):
    """Rango del eje Y ajustado a los datos y redondeado a un paso legible.

    Sin esto las marcas salen con valores como 471.778.915. El paso se elige de {1, 2, 2.5, 5}
    por potencia de diez para que queden ~6 divisiones: redondear a la potencia de diez pelada
    daría un paso tan grueso que el rango volvería a ser casi el mismo que el automático de
    Excel, y las series quedarían otra vez aplastadas contra un borde. El redondeo va hacia
    afuera, así que el rango nunca recorta un dato.
    """
    span = max(hi - lo, abs(hi) * 0.002, 1)
    lo -= span * margen_rel
    hi += span * margen_rel
    crudo = (hi - lo) / divisiones
    exp = 10 ** (len(str(int(crudo))) - 1) if crudo >= 1 else 1
    paso = next((m * exp for m in (1, 2, 2.5, 5, 10) if m * exp >= crudo), 10 * exp)
    return int(lo // paso * paso), int(-(-hi // paso) * paso)


def hoja_graficas(wb, data, por_medidor):
    """Dos gráficas: el modo de falla del banco B, y el caso sano, para contrastar.

    Las series se grafican desde un bloque de datos propio y no desde la hoja «Detalle»: ahí el
    centinela de saturación se escribe como texto («SATURADO») para que nadie lo promedie por
    error, y Excel no grafica texto — la serie del banco B desaparecería justo en el caso que
    hay que mostrar. Acá se escribe el número publicado, que es lo que hace visible la meseta.
    """
    ws = wb.create_sheet("Gráficas")
    anchos(ws, [102, 3, 18, 18, 18])
    titulo(ws, 1, "Cómo se ve el techo del banco B")

    # Un medidor donde el banco B ya saturó y otro donde no: el contraste ES el argumento.
    saturados = {m["medidor"] for m in data["medidores"]
                 for s in m["salud"] if s["reg"] == 40093 and s["inservible"]}
    elegidos = []
    for medidor in sorted(por_medidor):
        if medidor in saturados and not any(e[1] for e in elegidos):
            elegidos.append((medidor, True))
        elif medidor not in saturados and not any(not e[1] for e in elegidos):
            elegidos.append((medidor, False))

    col0 = 3          # el bloque de datos arranca en la columna C
    fila_datos = 3
    ancla = 3
    for medidor, saturado in elegidos:
        cab = fila_datos
        ws.cell(row=cab, column=col0, value="Hora").font = Font(bold=True, size=9, color=TINTA)
        ws.cell(row=cab, column=col0 + 1, value=f"{medidor} · 40232 (banco A)").font = Font(bold=True, size=9, color=TINTA)
        ws.cell(row=cab, column=col0 + 2, value=f"{medidor} · 40093 (banco B)").font = Font(bold=True, size=9, color=TINTA)
        fila = cab + 1
        valores = []
        for r in por_medidor[medidor]:
            if not r.get("ok"):
                continue
            a = (r["regs"].get("40232") or {}).get("value")
            b = (r["regs"].get("40093") or {}).get("value")
            ws.cell(row=fila, column=col0, value=a_bogota(r["ts"])).number_format = "hh:mm:ss"
            ws.cell(row=fila, column=col0 + 1, value=a).number_format = "#,##0"
            ws.cell(row=fila, column=col0 + 2, value=b).number_format = "#,##0"
            valores += [v for v in (a, b) if v is not None]
            fila += 1
        fin = fila - 1

        g = LineChart()
        g.title = (f"{medidor} — kWh rec: 40232 (banco A) vs 40093 (banco B)"
                   + (" · el banco B está clavado en su tope" if saturado else " · los dos bancos coinciden"))
        g.height, g.width = 8.5, 26
        for off, color in ((1, AZUL), (2, NARANJA)):
            g.add_data(Reference(ws, min_col=col0 + off, min_row=cab, max_row=fin), titles_from_data=True)
            linea(g.series[-1], color)
        g.set_categories(Reference(ws, min_col=col0, min_row=cab + 1, max_row=fin))
        estiliza_ejes(g, tick_skip=max(1, (fin - cab) // 10))
        # Rango ajustado a los datos. Estos son contadores acumulados de cientos de millones:
        # con el eje anclado en cero las dos series quedan aplastadas contra el borde superior y
        # no se distingue la que sube de la que está plana, que es justamente lo que hay que ver.
        # (Un eje truncado sería incorrecto en barras, donde el área codifica la magnitud; en una
        # línea de una serie sin cero natural, lo que codifica es la pendiente.)
        if valores:
            g.y_axis.scaling.min, g.y_axis.scaling.max = rango_redondeado(min(valores), max(valores))
        ws.add_chart(g, f"A{ancla}")
        ancla += 18
        fila_datos = fin + 3

    for txt in [
        "Cuando el acumulador pasa de 327.679.999 el banco B no da error: publica el centinela 0x7FFF/0 y "
        "se queda plano. Un consumidor que lo lea sin comprobar la saturación ve un contador congelado y no "
        "tiene forma de distinguirlo de una planta parada.",
        "En la gráfica del medidor saturado, la línea azul (banco A) sigue subiendo mientras la naranja "
        "(banco B) es una recta horizontal: los dos leen el mismo acumulador, pero uno ya no lo puede "
        "representar. Las columnas C–E son los datos de las gráficas.",
    ]:
        c = ws.cell(row=ancla, column=1, value=txt)
        c.font = Font(size=10, color=TINTA_2, italic=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ancla].height = 14 * max(2, -(-len(txt) // 100))
        ancla += 2
    return ws


def hoja_metodo(wb):
    ws = wb.create_sheet("Método")
    anchos(ws, [112])
    f = titulo(ws, 1, "Cómo se hizo esta comparación")
    bloques = [
        ("Qué se midió",
         "Los dos bancos completos (13 registros) más 40204 (kW tot) se leyeron sobre los 6 medidores cada "
         "5 s durante 1 h. Se guardaron las palabras de 16 bits CRUDAS además de los valores decodificados, "
         "para poder probar otra decodificación después sin volver a medir."),
        ("Mod10K no es INT32",
         "El banco B declara INT32-M10K: cada registro de 16 bits lleva 4 dígitos decimales, así que el "
         "valor es w0·10000 + w1. Leerlo como INT32 (w0·65536 + w1) da un número plausible y equivocado. Lo "
         "traicionero es que mientras w0 no cambia, el INCREMENTO de dos lecturas coincide con el correcto, "
         "así que el error solo aparece cuando w1 da la vuelta."),
        ("El orden de lectura se alterna a propósito",
         "Los dos bancos están en bloques Modbus distintos: son dos peticiones separadas por unos ms, y el "
         "contador sube entre una y otra. Con un orden fijo, esa diferencia tendría siempre el mismo signo y "
         "se podría confundir con un sesgo real. Alternando A→B y B→A, una diferencia por desfase cambia de "
         "signo con el orden y una discrepancia real no."),
        ("Tres evidencias independientes",
         "1) Concordancia instantánea: ¿los dos bancos publican el mismo número en cada tick?\n"
         "2) Identidades internas del banco B: del+rec debe ser del + rec. No depende de nada externo.\n"
         "3) Contraste contra la energía real: la integral de la potencia dice cuánta energía hubo."),
        ("Reproducir",
         "cd server && npm run shadow:registros           (captura 1 h, no toca producción)\n"
         "           npm run shadow:registros:analyze     (reporte y veredicto)\n"
         "           python scripts/export-registros-xlsx.py   (este archivo)"),
    ]
    f += 1
    for encabezado, cuerpo in bloques:
        c = ws.cell(row=f, column=1, value=encabezado)
        c.font = Font(bold=True, size=11, color=TINTA)
        f += 1
        c = ws.cell(row=f, column=1, value=cuerpo)
        c.font = Font(size=10, color=TINTA)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[f].height = 15 * (1 + len(cuerpo) // 100 + cuerpo.count("\n"))
        f += 2
    return ws


def main():
    salida = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        AQUI.parent.parent / "docs" / "analisis" / "comparacion-registros-energia.xlsx")
    data = cargar_resumen()
    por_medidor = cargar_jsonl()

    wb = Workbook()
    wb.remove(wb.active)
    hoja_veredicto(wb, data)
    hoja_lado_a_lado(wb, data)
    hoja_salud(wb, data)
    hoja_identidades(wb, data)
    hoja_energia_real(wb, data)
    hoja_detalle(wb, data, por_medidor)
    hoja_graficas(wb, data, por_medidor)
    hoja_metodo(wb)

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    print(f"Excel escrito: {salida}")
    print(f"  hojas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
