#!/usr/bin/env python3
"""Exporta a Excel la comparación 1-a-1 medidor primario vs medidor de reserva.

Las cifras del resumen y de los criterios NO se recalculan acá: salen de
`node scripts/analyze-reserva.js --json`, que es el mismo código que produce el reporte de
consola. Así el Excel y el reporte no pueden divergir. Este script solo da formato y agrega
las hojas de detalle tick a tick, que lee de los JSONL crudos.

Uso (desde server/):
    python scripts/export-reserva-xlsx.py [salida.xlsx]

Requiere `openpyxl` (pip install openpyxl) y `node` en el PATH.
"""
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent / "lib"))
from xlsxfmt import (  # noqa: E402
    AZUL, BOGOTA, BORDE_FINO, NARANJA, TINTA, TINTA_2,
    a_bogota, anchos, encabezados, estiliza_ejes, linea, nota, pct, pinta_estado, titulo,
)

AQUI = Path(__file__).resolve().parent
TRACES = AQUI.parent / "traces" / "reserva"
ANALIZADOR = AQUI / "analyze-reserva.js"


# ─── Carga de datos ───────────────────────────────────────────────────────────

def cargar_resumen():
    """Corre el analizador en modo JSON y devuelve su salida."""
    res = subprocess.run(
        ["node", str(ANALIZADOR), "--json"],
        capture_output=True, text=True, encoding="utf-8", cwd=str(AQUI.parent),
    )
    if res.returncode != 0:
        sys.exit(f"El analizador falló:\n{res.stderr}")
    return json.loads(res.stdout)


def cargar_jsonl():
    """Lee los JSONL crudos y agrupa por par: potencia y energía."""
    if not TRACES.is_dir():
        sys.exit(f"No existe {TRACES}. Corre primero: npm run shadow:reserva")
    por_par = {}
    for archivo in sorted(TRACES.glob("reserva-*.jsonl")):
        with archivo.open(encoding="utf-8") as fh:
            for linea_txt in fh:
                linea_txt = linea_txt.strip()
                if not linea_txt:
                    continue
                r = json.loads(linea_txt)
                b = por_par.setdefault(r["pair"], {"power": [], "energy": []})
                b["energy" if r.get("kind") == "energy" else "power"].append(r)
    for b in por_par.values():
        b["power"].sort(key=lambda r: r["ts"])
        b["energy"].sort(key=lambda r: r["ts"])
    return por_par


# ─── Hojas ────────────────────────────────────────────────────────────────────

def hoja_resumen(wb, data):
    ws = wb.create_sheet("Resumen")
    anchos(ws, [26, 17, 17, 15, 13, 15, 13, 15, 15, 13, 13, 34])
    v = data["ventana"]
    f = titulo(ws, 1, "Comparación 1-a-1: medidores de reserva vs. primarios (Modbus TCP)", 15)
    f = nota(ws, f, f"Ventana {a_bogota(v['desde']):%Y-%m-%d %H:%M:%S} → {a_bogota(v['hasta']):%H:%M:%S} "
                    f"hora Bogotá ({v['horas']:.2f} h) · {data['registros']} registros · "
                    f"generado {datetime.now(BOGOTA):%Y-%m-%d %H:%M} Bogotá", cols=12)
    f += 1

    f = titulo(ws, f, "Resultado por par", 12)
    f = encabezados(ws, f, [
        "Bloque", "Primario", "Reserva", "|P| media (kW)", "Sesgo (kW)", "Sesgo (%)",
        "p95 |Δ| (kW)", "p95 |Δ| (%)", "Energía 1 h: dif (%)", "Nulls prim.", "Nulls res.", "Veredicto",
    ])
    primera = f
    for b in data["bloques"]:
        s, e, h = b["s"], b["e"], b["hosts"]
        # El acuerdo se lee de los ticks limpios cuando hay suficientes, igual que el reporte.
        ac = s["clean"] if s["clean"]["n"] >= data["umbrales"]["minCleanTicks"] else s["all"]
        etiqueta = "UNIDAD GEC3 (suma)" if b["pair"] == "UNIDAD_GEC3" else b["pair"]
        vals = [
            etiqueta,
            h["prim"] if h else "—", h["res"] if h else "—",
            s["meanAbsPower"], ac["bias"], pct(ac["biasPct"]),
            ac["absP95"], pct(ac["absP95Pct"]),
            pct(e["relDiffPct"]) if e and e.get("relDiffPct") is not None else None,
            s["primNull"], s["resNull"], b["ev"]["verdict"],
        ]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=f, column=i, value=val)
            c.border = BORDE_FINO
            if i == 1:
                c.font = Font(bold=True, size=10)
            elif i == 4:
                c.number_format = "#,##0"
            elif i in (5, 7):
                c.number_format = "#,##0.000"
            elif i in (6, 8, 9):
                c.number_format = "0.000%"
            elif i == 12:
                pinta_estado(c, b["ev"]["verdict"].startswith("INTERCAMBIABLE"))
        f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)

    f += 1
    f = titulo(ws, f, "Lectura", 12)
    for linea in [
        "· Disponibilidad perfecta: 0 nulls en los 6 medidores durante toda la hora.",
        "· El sesgo cambia de signo entre pares, que es la firma de ruido de medición independiente "
        "y no de un error de calibración sistemático (un error de escala apuntaría al mismo lado en todos).",
        "· La energía acumulada es la evidencia más fuerte porque no depende del instante de muestreo: "
        "el contador de cada lado avanzó lo mismo en la hora.",
        "· El valor que consume el dashboard es GEC3 como suma de sus 2 medidores, y ahí el sesgo es 0.000 %: "
        "sumar promedia el ruido independiente de los dos pares.",
        "· Conclusión: la reserva sirve como fallback REAL, con instrumentos físicamente independientes, "
        "a diferencia del PME, que lee los mismos ION8650 y sirve un valor congelado cuando estos se caen.",
    ]:
        c = ws.cell(row=f, column=1, value=linea)
        c.font = Font(size=10, color=TINTA)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=12)
        ws.row_dimensions[f].height = 28
        f += 1

    f += 1
    p = data.get("pareo")
    if p:
        f = titulo(ws, f, "Verificación del pareo de GEC3", 12)
        f = nota(ws, f, "Que .5↔.7 y .6↔.8 era una suposición por el orden de las IPs. Se comprueba "
                        "recalculando el acuerdo con el mapeo cruzado: gana el que dé menor Δ.", cols=12)
        for etiqueta, val in [
            (f"Mapeo asumido  ({p['hosts']['p1']}↔{p['hosts']['r1']}, {p['hosts']['p2']}↔{p['hosts']['r2']})", p["errAsumido"]),
            (f"Mapeo cruzado  ({p['hosts']['p1']}↔{p['hosts']['r2']}, {p['hosts']['p2']}↔{p['hosts']['r1']})", p["errCruzado"]),
        ]:
            # Se fusionan A:C porque la etiqueta lleva las 4 IPs y si no queda cortada.
            ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
            ws.cell(row=f, column=1, value=etiqueta).font = Font(size=10)
            c = ws.cell(row=f, column=4, value=val)
            c.number_format = "#,##0.0"
            ws.cell(row=f, column=5, value="kW  (|Δ| p50 sumado)").font = Font(size=10, color=TINTA_2)
            f += 1
        c = ws.cell(row=f, column=1, value={
            "ASUMIDO": f"✓ CONFIRMADO: el mapeo asumido ajusta mejor por {p['margen']:.1f} kW.",
            "CRUZADO": f"✗ Las IPs de reserva están al revés: el mapeo cruzado ajusta mejor por {p['margen']:.1f} kW.",
        }.get(p["veredicto"], "~ Los dos mapeos ajustan casi igual: el pareo no se puede resolver por valor."))
        pinta_estado(c, p["veredicto"] == "ASUMIDO")
    return ws


def hoja_criterios(wb, data):
    ws = wb.create_sheet("Criterios")
    anchos(ws, [22, 8, 40, 15, 78])
    u = data["umbrales"]
    f = titulo(ws, 1, "Criterios de éxito, bloque por bloque")
    f = nota(ws, f, "Las tolerancias de acuerdo son relativas a la potencia media absoluta de la ventana, "
                    "no absolutas en kW: a 224 MW un umbral fijo sería trivial de pasar, y con la unidad en "
                    "reserva (~0.7 MW) el error relativo punto a punto se dispara sin que el medidor tenga nada malo.", cols=5)
    f = nota(ws, f, f"Umbrales: sesgo ≤{u['biasPctMax']}% · p95 ≤{u['p95PctMax']}% · "
                    f"pendiente {u['slopeMin']}–{u['slopeMax']} · offset ≤{u['interceptPctMax']}% · "
                    f"energía ≤{u['energyRelPctMax']}% · nulls ≤{u['resNullRateMax']}% · p99 <{u['timeoutMs']}ms", cols=5)
    f += 1
    f = encabezados(ws, f, ["Bloque", "Criterio", "Qué exige", "Resultado", "Medido"])
    primera = f
    for b in data["bloques"]:
        etiqueta = "UNIDAD GEC3 (suma)" if b["pair"] == "UNIDAD_GEC3" else b["pair"]
        for clave, crit in b["ev"]["c"].items():
            ok = crit["ok"]
            fila = [etiqueta, clave.upper(), crit["label"],
                    {True: "✓ CUMPLE", False: "✗ NO CUMPLE"}.get(ok, "· SIN SEÑAL"), crit["detail"]]
            for i, val in enumerate(fila, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.border = BORDE_FINO
                c.font = Font(size=10, color=TINTA)
                c.alignment = Alignment(vertical="top", wrap_text=(i == 5))
                if i == 4:
                    pinta_estado(c, ok)
                    c.alignment = Alignment(horizontal="center", vertical="top")
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)

    f += 1
    f = nota(ws, f, "«Sin señal» no reprueba: es un criterio que la ventana no permitió medir. Se informa "
                    "para que un ✓ no se lea como respaldado por datos cuando no lo está.", cols=5)
    nota(ws, f, "R² se imprime como contexto pero no decide: R² = 1 − SSE/SST, y SST depende de cuánto se "
                "movió la planta, no de la calidad del medidor. Dos medidores idénticos sobre una planta "
                "constante darían R² ≈ 0. La dispersión ya la mide C3, correctamente normalizada.", cols=5)
    return ws


COLS_DETALLE = ["Hora (Bogotá)", "Primario (kW)", "Reserva (kW)", "Δ = prim − res (kW)",
                "Δ relativo", "Desfase (ms)", "Lat. prim (ms)", "Lat. res (ms)", "Estado"]


def hoja_detalle(wb, nombre, filas, subtitulo):
    """Una fila por tick, más dos gráficas: la superposición y el residuo."""
    ws = wb.create_sheet(nombre)
    anchos(ws, [18, 16, 16, 19, 13, 13, 14, 13, 22])
    f = titulo(ws, 1, f"Detalle tick a tick — {nombre}")
    f = nota(ws, f, subtitulo, cols=9)
    f += 1
    cab = f
    f = encabezados(ws, f, COLS_DETALLE)
    inicio = f

    for r in filas:
        vals = [r["hora"], r["prim"], r["res"], r["diff"], pct(r["relPct"]),
                r["skew"], r["latPrim"], r["latRes"], r["estado"]]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=f, column=i, value=val)
            c.font = Font(size=9, color=TINTA)
            if i == 1:
                c.number_format = "hh:mm:ss"
            elif i in (2, 3, 4):
                c.number_format = "#,##0.000"
            elif i == 5:
                c.number_format = "0.0000%"
            if i == 9 and r["estado"] != "ambos OK":
                c.font = Font(size=9, bold=True, color="8E2020")
        f += 1
    fin = f - 1
    ws.freeze_panes = ws.cell(row=inicio, column=1)
    ws.auto_filter.ref = f"A{cab}:I{fin}"

    horas = Reference(ws, min_col=1, min_row=inicio, max_row=fin)

    # Gráfica 1 — superposición. Las dos curvas se tapan entre sí: ese ES el resultado.
    g1 = LineChart()
    # Unidades y zona horaria van EN EL TÍTULO: con el eje X abajo, un título de eje se
    # imprime a la misma altura que las etiquetas de hora y se montan entre sí.
    g1.title = f"{nombre} — potencia leída por cada medidor (kW, hora Bogotá)"
    g1.height, g1.width = 8.5, 30
    for col, color, etiqueta in ((2, AZUL, "Primario"), (3, NARANJA, "Reserva")):
        ref = Reference(ws, min_col=col, min_row=cab, max_row=fin)
        g1.add_data(ref, titles_from_data=True)
        s = linea(g1.series[-1], color)
        s.tx.strRef.f = f"'{nombre}'!${get_column_letter(col)}${cab}"
        del etiqueta
    g1.set_categories(horas)
    estiliza_ejes(g1)
    ws.add_chart(g1, "K3")

    # Gráfica 2 — el residuo, que es donde de verdad se ve el desacuerdo.
    g2 = LineChart()
    g2.title = f"{nombre} — diferencia entre los dos medidores, Δ = primario − reserva (kW, hora Bogotá)"
    g2.height, g2.width = 8.5, 30
    g2.add_data(Reference(ws, min_col=4, min_row=cab, max_row=fin), titles_from_data=True)
    linea(g2.series[0], AZUL)
    g2.set_categories(horas)
    estiliza_ejes(g2)
    g2.legend = None  # una sola serie: el título ya la nombra
    ws.add_chart(g2, "K21")
    return ws


def hoja_energia(wb, por_par, data):
    ws = wb.create_sheet("Energía")
    anchos(ws, [18, 14, 20, 20, 20, 20, 16])
    f = titulo(ws, 1, "Contador de energía acumulada (registro 40232, kWh)")
    f = nota(ws, f, "Los totales absolutos NO son comparables: cada medidor arrancó su cuenta en un momento "
                    "distinto (se vio 460.001.152 contra 8.054.456 en el mismo par). Lo único comparable es "
                    "cuánto AVANZÓ cada uno durante la ventana — las dos últimas columnas.", cols=7)
    f = nota(ws, f, "Es la evidencia más fuerte de la comparación porque no depende del instante de muestreo, "
                    "a diferencia de la potencia instantánea.", cols=7)
    f += 1
    f = encabezados(ws, f, ["Hora (Bogotá)", "Par", "Primario (kWh)", "Reserva (kWh)",
                            "Avance primario (kWh)", "Avance reserva (kWh)", "Diferencia"])
    primera = f
    for par in sorted(por_par):
        recs = [r for r in por_par[par]["energy"] if r["prim"]["ok"] and r["res"]["ok"]]
        if not recs:
            continue
        p0, r0 = recs[0]["prim"]["raw"], recs[0]["res"]["raw"]
        for r in recs:
            dp, dr = r["prim"]["raw"] - p0, r["res"]["raw"] - r0
            vals = [a_bogota(r["ts"]), par, r["prim"]["raw"], r["res"]["raw"], dp, dr,
                    (abs(dr - dp) / abs(dp)) if dp else None]
            for i, val in enumerate(vals, start=1):
                c = ws.cell(row=f, column=i, value=val)
                c.font = Font(size=9, color=TINTA)
                if i == 1:
                    c.number_format = "hh:mm:ss"
                elif i in (3, 4, 5, 6):
                    c.number_format = "#,##0"
                elif i == 7:
                    c.number_format = "0.000%"
            f += 1
    ws.freeze_panes = ws.cell(row=primera, column=1)

    f += 1
    f = titulo(ws, f, "Avance total de la ventana", 12)
    f = encabezados(ws, f, ["Bloque", "", "Avance primario (kWh)", "Avance reserva (kWh)", "Diferencia (kWh)", "Diferencia (%)"])
    for b in data["bloques"]:
        e = b["e"]
        if not e or e["prim"]["delta"] is None:
            continue
        etiqueta = "UNIDAD GEC3 (suma)" if b["pair"] == "UNIDAD_GEC3" else b["pair"]
        vals = [etiqueta, "", e["prim"]["delta"], e["res"]["delta"],
                e["res"]["delta"] - e["prim"]["delta"], pct(e["relDiffPct"])]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=f, column=i, value=val)
            c.border = BORDE_FINO
            c.font = Font(size=10, bold=(i == 1))
            if i in (3, 4, 5):
                c.number_format = "#,##0"
            elif i == 6:
                c.number_format = "0.000%"
        f += 1
    return ws


def hoja_metodo(wb, data):
    ws = wb.create_sheet("Método")
    anchos(ws, [110])
    f = titulo(ws, 1, "Cómo se hizo esta medición")
    bloques = [
        ("Qué se comparó",
         "Cada par (primario, reserva) se leyó por Modbus TCP en paralelo cada 2 s durante 1 h, registro 40204 "
         "(kW total, INT32, word order high, escala /1000, unitId 1). En paralelo, cada 60 s se leyó el contador "
         "de energía acumulada del registro 40232."),
        ("Por qué el Δ va con signo",
         "Δ = primario − reserva se guarda con signo. Un sesgo sistemático es calibración y una dispersión "
         "simétrica es ruido: un promedio de valores absolutos los confundiría."),
        ("Por qué las tolerancias son relativas",
         "Se normalizan contra la potencia media absoluta de la ventana. A 224 MW un umbral fijo en kW sería "
         "trivial de pasar; con la unidad en reserva (~0.7 MW) el error relativo punto a punto se dispara sin "
         "que el medidor tenga nada malo."),
        ("Qué ticks se descartan",
         "Los dos medidores de un par no se leen en el mismo instante. Ese desfase produce una diferencia "
         "aparente de |dP/dt| · skew que no es del medidor. Se descarta un tick solo cuando ese error alcanza "
         "para explicar parte apreciable de la tolerancia — NO por el mero hecho de que haya rampa: a 224 MW la "
         "potencia fluctúa decenas de kW entre muestras de 2 s, así que descartar por rampa tiraría la ventana "
         "entera, y encima es la rampa la que le da rango a la regresión."),
        ("Por qué la regresión lleva ± su error estándar",
         "reserva = a · primario + b separa un error de escala (a ≠ 1, p. ej. relación de TC/TP distinta) de un "
         "offset constante (b ≠ 0). Pero si la unidad se movió poco, «a» sale lejos de 1 con una incertidumbre "
         "enorme y el número se lee igual de contundente. Con ±2·error estándar se puede decir «0.971 ± 0.010» y "
         "concluir que la ventana no alcanza para juzgar la escala, en vez de reprobar un medidor sano."),
        ("Por qué R² no decide",
         "R² = 1 − SSE/SST, y SST depende de cuánto se movió la planta, no de la calidad del medidor: dos "
         "medidores idénticos sobre una planta perfectamente constante darían R² ≈ 0. Además es redundante con "
         "el criterio de dispersión, que ya está correctamente normalizado. Se imprime como contexto."),
        ("El registro de energía es 40232, no 40230",
         "La documentación del repo decía 40230. Ese registro es kWh DELIVERED y en Gecelca no se mueve: la "
         "frontera es de entrada, así que con la planta generando lo que avanza es 40232 (kWh RECEIVED). "
         "Verificado contra la integral de la potencia en los 6 medidores. El mapa completo de los dos bancos "
         "de energía del ION8650 está en docs/analisis/comparacion-registros-energia.md."),
        ("Reproducir",
         "cd server && npm run shadow:reserva  (captura 1 h, no toca producción)\n"
         "           npm run shadow:reserva:analyze  (reporte y veredicto)\n"
         "           python scripts/export-reserva-xlsx.py  (este archivo)\n"
         "Detalle en docs/analisis/comparacion-medidores-reserva.md"),
    ]
    f += 1
    for encabezado, cuerpo in bloques:
        c = ws.cell(row=f, column=1, value=encabezado)
        c.font = Font(bold=True, size=11, color=TINTA)
        f += 1
        c = ws.cell(row=f, column=1, value=cuerpo)
        c.font = Font(size=10, color=TINTA)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[f].height = 15 * (1 + len(cuerpo) // 95 + cuerpo.count("\n"))
        f += 2
    del data
    return ws


# ─── Armado de las filas de detalle ───────────────────────────────────────────

def filas_par(recs):
    out = []
    for r in recs:
        ok = bool(r["prim"]["ok"] and r["res"]["ok"])
        estado = "ambos OK" if ok else (
            "primario NULL" if not r["prim"]["ok"] and r["res"]["ok"] else
            "reserva NULL" if r["prim"]["ok"] else "ambos NULL")
        out.append({
            "hora": a_bogota(r["ts"]),
            "prim": r["prim"]["kw"], "res": r["res"]["kw"],
            "diff": r.get("diffKw"), "relPct": r.get("relDiffPct"),
            "skew": r.get("skewMs"),
            "latPrim": r["prim"]["latencyMs"], "latRes": r["res"]["latencyMs"],
            "estado": estado,
        })
    return out


def filas_unidad(por_par):
    """GEC3 como suma de sus 2 medidores: es el valor que publica el dashboard."""
    a = {r["ts"]: r for r in por_par.get("GEC3_1", {}).get("power", [])}
    b = {r["ts"]: r for r in por_par.get("GEC3_2", {}).get("power", [])}
    out = []
    for ts in sorted(a.keys() & b.keys()):
        ra, rb = a[ts], b[ts]
        if not all((ra["prim"]["ok"], ra["res"]["ok"], rb["prim"]["ok"], rb["res"]["ok"])):
            continue
        prim = ra["prim"]["kw"] + rb["prim"]["kw"]
        res = ra["res"]["kw"] + rb["res"]["kw"]
        diff = prim - res
        out.append({
            "hora": a_bogota(ts), "prim": prim, "res": res, "diff": diff,
            "relPct": (abs(diff) / abs(prim) * 100) if prim else None,
            "skew": max(ra.get("skewMs") or 0, rb.get("skewMs") or 0),
            "latPrim": None, "latRes": None, "estado": "ambos OK",
        })
    return out


def main():
    salida = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        AQUI.parent.parent / "docs" / "analisis" / "comparacion-medidores-reserva.xlsx")
    data = cargar_resumen()
    por_par = cargar_jsonl()

    wb = Workbook()
    wb.remove(wb.active)
    hoja_resumen(wb, data)
    hoja_criterios(wb, data)
    for par in sorted(p for p in por_par if por_par[p]["power"]):
        h = next((b["hosts"] for b in data["bloques"] if b["pair"] == par), None)
        sub = f"Primario {h['prim']} vs reserva {h['res']} · 1 lectura cada 2 s" if h else ""
        hoja_detalle(wb, par, filas_par(por_par[par]["power"]), sub)
    unidad = filas_unidad(por_par)
    if unidad:
        hoja_detalle(wb, "UNIDAD GEC3", unidad,
                     "Suma de los 2 medidores de GEC3 — el valor que publica el dashboard (combine:'sum'). "
                     "Es inmune a que el pareo entre medidores estuviera cruzado.")
    hoja_energia(wb, por_par, data)
    hoja_metodo(wb, data)

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    print(f"Excel escrito: {salida}")
    print(f"  hojas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
